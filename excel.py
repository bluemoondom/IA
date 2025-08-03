# -*- coding: utf-8 -*-
"""
Created on Fri Oct  4 14:06:18 2024

@author: dominika
"""



from openpyxl import load_workbook

myfilename = "C:/Users/pille/Desktop/excel/vzor_1.xlsx"
mynewfilename = "C:/Users/pille/Desktop/excel/my_vzor_1.xlsx"
mystring = "Pol1557"
wb = load_workbook(myfilename)
ws = wb.active

for row in ws.iter_rows(min_col=1, max_col=7, min_row=170, max_row=10000):
    if row[5].value == mystring:
        for cell in row:
            if cell.value == mystring:
                mycolumn = cell.column_letter
                myrow = cell.row
                mysheet = cell.parent
                print(mycolumn)
                print(myrow)
                print(mysheet)
            #print(cell.column_letter, end=" ")    
        #print()
        

workbook = load_workbook(myfilename)
 

mysheetstr = str(mysheet).replace("<Worksheet \"", "").replace("\">", "")
workbook.active = workbook[mysheetstr]
sheet = workbook.active

columnrow = "J" + str(myrow)
print(columnrow)
sheet[columnrow] = "1,5"

workbook.save(mynewfilename)